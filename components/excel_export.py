import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from data.transforms import compute_entity_table, compute_metrics, count_unique_providers


blue_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
light_blue = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True, size=10)
bold_font = Font(bold=True, size=10)
header_font = Font(bold=True, size=12, color="1A1A2E")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _safe(text):
    return str(text).encode("latin-1", errors="replace").decode("latin-1")


def _write_table(ws, data, start_row, table_name):
    """Write a DataFrame as a formatted Excel Table with filters."""
    if data.empty:
        return start_row

    headers = list(data.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=_safe(h))
        cell.font = white_font
        cell.fill = blue_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for ri, (_, row) in enumerate(data.iterrows()):
        for ci, h in enumerate(headers, 1):
            val = row[h]
            cell = ws.cell(row=start_row + 1 + ri, column=ci, value=_safe(val) if isinstance(val, str) else val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center") if ci > 1 else Alignment()
            if ri % 2 == 0:
                cell.fill = light_blue

    end_row = start_row + len(data)
    end_col = get_column_letter(len(headers))
    ref = f"A{start_row}:{end_col}{end_row}"
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tab)

    return end_row + 2


def generate_full_export(df, period_col):
    """Generate a multi-sheet Excel workbook with all dashboard data."""
    wb = Workbook()
    is_weekly = period_col == "week_of"

    # ==================== Sheet 1: Raw Data ====================
    ws1 = wb.active
    ws1.title = "Referral Data"

    raw_cols = [
        "REFERRAL_DATE", "REFERRAL_ID", "PARTNER_ASSIGNMENT", "PPM", "TEAM_TYPE",
        "REFERRING_CLINIC", "REFERRING_CLINIC_ZIP", "provider_id",
        "REFERRING_PHYSICIAN", "REFERRAL_SOURCE_TYPE", "PATIENT_INSURANCE_NAME",
        "intake_started", "visit_booked", "visit_completed",
        "INTAKE_ACTION_STATUS", "TERMINATION_REASON",
        "APPOINTMENT_SOURCE_FIRST_SCHEDULED",
    ]
    raw_cols = [c for c in raw_cols if c in df.columns]
    raw_df = df[raw_cols].sort_values("REFERRAL_DATE", ascending=False).copy()
    raw_df["REFERRAL_DATE"] = raw_df["REFERRAL_DATE"].dt.strftime("%Y-%m-%d")

    _write_table(ws1, raw_df, 1, "ReferralData")
    for i in range(1, len(raw_cols) + 1):
        ws1.column_dimensions[get_column_letter(i)].width = 18

    # ==================== Sheet 2: Account Summary ====================
    ws2 = wb.create_sheet("Account Summary")

    acct_table = compute_entity_table(df, "PARTNER_ASSIGNMENT", period_col)
    if "TEAM_TYPE" in df.columns:
        team_map = df.groupby("PARTNER_ASSIGNMENT")["TEAM_TYPE"].agg(
            lambda x: x.mode().iloc[0] if len(x.mode()) > 0 else ""
        ).reset_index()
        team_map.columns = ["PARTNER_ASSIGNMENT", "TEAM_TYPE"]
        acct_table = acct_table.merge(team_map, on="PARTNER_ASSIGNMENT", how="left")

    acct_display = acct_table[["PARTNER_ASSIGNMENT"] +
        (["TEAM_TYPE"] if "TEAM_TYPE" in acct_table.columns else []) +
        ["category", "referrals", "days_since_last", "pct_intake", "pct_booked", "pct_completed", "trend"]
    ].copy()
    acct_display["pct_intake"] = (acct_display["pct_intake"] * 100).round(1)
    acct_display["pct_booked"] = (acct_display["pct_booked"] * 100).round(1)
    acct_display["pct_completed"] = (acct_display["pct_completed"] * 100).round(1)
    acct_display["trend"] = (acct_display["trend"] * 100).round(0)
    acct_display["days_since_last"] = acct_display["days_since_last"].fillna(0).astype(int)
    acct_display = acct_display.rename(columns={
        "PARTNER_ASSIGNMENT": "Account", "TEAM_TYPE": "Team Type",
        "category": "Status", "referrals": "Referrals",
        "days_since_last": "Days Silent", "pct_intake": "% Intake",
        "pct_booked": "% Booked", "pct_completed": "% Completed", "trend": "Trend %",
    })

    _write_table(ws2, acct_display, 1, "AccountSummary")
    ws2.column_dimensions["A"].width = 30
    for i in range(2, 11):
        ws2.column_dimensions[get_column_letter(i)].width = 14

    # ==================== Sheet 3: PPM Summary ====================
    ws3 = wb.create_sheet("PPM Summary")

    ppm_rows = []
    for ppm in sorted(df["PPM"].unique()):
        ppm_df = df[df["PPM"] == ppm]
        m = compute_metrics(ppm_df)
        accounts = ppm_df["PARTNER_ASSIGNMENT"].nunique()
        clinics = ppm_df["REFERRING_CLINIC"].nunique()
        ppm_rows.append({
            "PPM": ppm,
            "Accounts": accounts,
            "Clinics": clinics,
            "Referrals": m["referrals"],
            "Unique Providers": m["unique_providers"],
            "% Intake": round(m["pct_intake"] * 100, 1),
            "% Booked": round(m["pct_booked"] * 100, 1),
            "% Completed": round(m["pct_completed"] * 100, 1),
        })
    ppm_summary = pd.DataFrame(ppm_rows).sort_values("Referrals", ascending=False)

    _write_table(ws3, ppm_summary, 1, "PPMSummary")
    ws3.column_dimensions["A"].width = 22
    for i in range(2, 9):
        ws3.column_dimensions[get_column_letter(i)].width = 16

    # ==================== Sheet 4: Monthly Pivot ====================
    ws4 = wb.create_sheet("Monthly by Account")

    monthly = df.groupby(["PARTNER_ASSIGNMENT", "month_of"]).agg(
        referrals=("REFERRAL_ID", "count"),
    ).reset_index()
    monthly["month_of"] = monthly["month_of"].astype(str)
    pivot = monthly.pivot(index="PARTNER_ASSIGNMENT", columns="month_of", values="referrals").fillna(0).astype(int)
    pivot = pivot.reset_index().rename(columns={"PARTNER_ASSIGNMENT": "Account"})
    pivot["Total"] = pivot.iloc[:, 1:].sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False)

    _write_table(ws4, pivot, 1, "MonthlyByAccount")
    ws4.column_dimensions["A"].width = 30
    for i in range(2, len(pivot.columns) + 1):
        ws4.column_dimensions[get_column_letter(i)].width = 12

    # ==================== Sheet 5: Clinic Rankings ====================
    ws5 = wb.create_sheet("Clinic Rankings")

    clinic_table = compute_entity_table(df, "REFERRING_CLINIC", period_col, include_account=True)
    clinic_display = clinic_table[["REFERRING_CLINIC", "PARTNER_ASSIGNMENT", "category",
        "referrals", "days_since_last", "pct_intake", "pct_booked", "pct_completed", "trend"]].copy()
    clinic_display["pct_intake"] = (clinic_display["pct_intake"] * 100).round(1)
    clinic_display["pct_booked"] = (clinic_display["pct_booked"] * 100).round(1)
    clinic_display["pct_completed"] = (clinic_display["pct_completed"] * 100).round(1)
    clinic_display["trend"] = (clinic_display["trend"] * 100).round(0)
    clinic_display["days_since_last"] = clinic_display["days_since_last"].fillna(0).astype(int)
    clinic_display = clinic_display.rename(columns={
        "REFERRING_CLINIC": "Clinic", "PARTNER_ASSIGNMENT": "Account",
        "category": "Status", "referrals": "Referrals",
        "days_since_last": "Days Silent", "pct_intake": "% Intake",
        "pct_booked": "% Booked", "pct_completed": "% Completed", "trend": "Trend %",
    })

    _write_table(ws5, clinic_display, 1, "ClinicRankings")
    ws5.column_dimensions["A"].width = 35
    ws5.column_dimensions["B"].width = 25
    for i in range(3, 10):
        ws5.column_dimensions[get_column_letter(i)].width = 14

    # ==================== Sheet 6: Provider Rankings ====================
    ws6 = wb.create_sheet("Provider Rankings")

    prov_table = compute_entity_table(df, "REFERRING_PHYSICIAN", period_col, include_account=True)
    prov_display = prov_table[["REFERRING_PHYSICIAN", "PARTNER_ASSIGNMENT", "category",
        "referrals", "days_since_last", "pct_intake", "pct_booked", "pct_completed", "trend"]].copy()
    prov_display["pct_intake"] = (prov_display["pct_intake"] * 100).round(1)
    prov_display["pct_booked"] = (prov_display["pct_booked"] * 100).round(1)
    prov_display["pct_completed"] = (prov_display["pct_completed"] * 100).round(1)
    prov_display["trend"] = (prov_display["trend"] * 100).round(0)
    prov_display["days_since_last"] = prov_display["days_since_last"].fillna(0).astype(int)
    prov_display = prov_display.rename(columns={
        "REFERRING_PHYSICIAN": "Provider", "PARTNER_ASSIGNMENT": "Account",
        "category": "Status", "referrals": "Referrals",
        "days_since_last": "Days Silent", "pct_intake": "% Intake",
        "pct_booked": "% Booked", "pct_completed": "% Completed", "trend": "Trend %",
    })

    _write_table(ws6, prov_display, 1, "ProviderRankings")
    ws6.column_dimensions["A"].width = 25
    ws6.column_dimensions["B"].width = 25
    for i in range(3, 10):
        ws6.column_dimensions[get_column_letter(i)].width = 14

    # ==================== Sheet 7: Action Plan ====================
    ws7 = wb.create_sheet("Action Plans")

    periods = sorted(df[period_col].dropna().unique())
    if len(periods) >= 2:
        curr_df = df[df[period_col] == periods[-1]]
        prev_df = df[df[period_col] == periods[-2]]
        all_prior = df[df[period_col] < periods[-1]]

        action_rows = []
        for ppm in sorted(df["PPM"].unique()):
            ppm_curr = curr_df[curr_df["PPM"] == ppm]
            ppm_prev = prev_df[prev_df["PPM"] == ppm]
            ppm_prior = all_prior[all_prior["PPM"] == ppm]

            curr_provs = set(ppm_curr["provider_id"].dropna())
            prev_provs = set(ppm_prev["provider_id"].dropna())
            all_prior_provs = set(ppm_prior["provider_id"].dropna())

            # New providers
            for pid in curr_provs - all_prior_provs:
                sub = ppm_curr[ppm_curr["provider_id"] == pid]
                name = sub["REFERRING_PHYSICIAN"].iloc[0] if not sub.empty else pid
                acct = sub["PARTNER_ASSIGNMENT"].iloc[0] if not sub.empty else ""
                action_rows.append({"PPM": ppm, "Action": "Say Thank You", "Type": "Provider",
                    "Name": _safe(name), "Account": _safe(acct), "Referrals": len(sub)})

            # Dropped providers
            for pid in prev_provs - curr_provs:
                sub = ppm_prev[ppm_prev["provider_id"] == pid]
                name = sub["REFERRING_PHYSICIAN"].iloc[0] if not sub.empty else pid
                acct = sub["PARTNER_ASSIGNMENT"].iloc[0] if not sub.empty else ""
                action_rows.append({"PPM": ppm, "Action": "Re-engage", "Type": "Provider",
                    "Name": _safe(name), "Account": _safe(acct), "Referrals": len(sub)})

        if action_rows:
            action_df = pd.DataFrame(action_rows).sort_values(["PPM", "Action", "Referrals"], ascending=[True, True, False])
            _write_table(ws7, action_df, 1, "ActionPlans")
            ws7.column_dimensions["A"].width = 20
            ws7.column_dimensions["B"].width = 16
            ws7.column_dimensions["C"].width = 10
            ws7.column_dimensions["D"].width = 25
            ws7.column_dimensions["E"].width = 25
            ws7.column_dimensions["F"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
