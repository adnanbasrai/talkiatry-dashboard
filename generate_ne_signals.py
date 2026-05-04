"""
generate_ne_signals.py
----------------------
Standalone script — reads NE Control Tower CSV and outputs NE_Account_Signals.xlsx.

Signals per account (one row per account):
  1. Referral Trend       — refs/working day MoM % change → percentile → STRONG/GROWING/FLAT/DECLINING/AT RISK
  2. Ref→Intake Current   — March rate vs absolute thresholds → HEALTHY/WATCH/AT RISK
  3. Intake→Booked Curr   — March rate vs absolute thresholds → HEALTHY/WATCH/AT RISK
  4. Ref→Intake MoM       — March vs Feb pp change → percentile → STRONG/GROWING/FLAT/DECLINING/AT RISK
  5. Intake→Booked MoM    — March vs Feb pp change → percentile → STRONG/GROWING/FLAT/DECLINING/AT RISK
  6. M1 Retention Current — Feb-new providers retained in March → absolute thresholds → STRONG/MODERATE/LOW
  7. M1 Retention MoM     — Feb cohort vs Jan cohort pp change → percentile → STRONG/GROWING/FLAT/DECLINING/AT RISK

Run:  python3 generate_ne_signals.py
Output: NE_Account_Signals.xlsx
"""

from __future__ import annotations

import numpy as np
import pandas as pd
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
DATA_PATH   = BASE_DIR / "data" / "NE_control_tower.csv"
OUTPUT_PATH = BASE_DIR / "NE_Account_Signals.xlsx"

# ── Date config ────────────────────────────────────────────────────────────────
TODAY        = date(2026, 4, 24)
CURR_MONTH   = "2026-04"   # in-flight
LAST_MONTH   = "2026-03"   # last complete  (conversion current + M1 "current" retention window)
PRIOR_MONTH  = "2026-02"   # prior complete (conversion MoM base + M1 "current" cohort)
OLDEST_MONTH = "2026-01"   # for M1 MoM prior cohort

# Working days
_APR_START = "2026-04-01"
_APR_END   = (TODAY + timedelta(days=1)).isoformat()          # inclusive of today
_MAR_START = "2026-03-01"
_MAR_END   = "2026-04-01"

APR_WORKDAYS = int(np.busday_count(_APR_START, _APR_END))    # elapsed through today
MAR_WORKDAYS = int(np.busday_count(_MAR_START, _MAR_END))    # full March

# ── Thresholds ─────────────────────────────────────────────────────────────────
# Percentile tiers (used for referral trend, conversion MoM, M1 MoM)
PCT_TIERS = [
    ("STRONG",    80),
    ("GROWING",   60),
    ("FLAT",      40),
    ("DECLINING", 20),
    ("AT RISK",    0),
]

# Conversion absolute thresholds (current rate)
INTAKE_ABS  = {"HEALTHY": 0.55, "WATCH": 0.45}   # < 0.45 → AT RISK
BOOKED_ABS  = {"HEALTHY": 0.70, "WATCH": 0.60}   # < 0.60 → AT RISK

# M1 absolute thresholds (current rate)
M1_ABS = {"STRONG": 0.35, "MODERATE": 0.25}       # < 0.25 → LOW

MIN_REFS    = 5   # minimum refs for conversion to be meaningful
MIN_COHORT  = 3   # minimum new providers for M1 to be meaningful
NA_LABEL    = "Not enough referrals"


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def pct_status(percentile: float | None) -> str:
    if percentile is None or (isinstance(percentile, float) and np.isnan(percentile)):
        return NA_LABEL
    for label, threshold in PCT_TIERS:
        if percentile >= threshold:
            return label
    return "AT RISK"


def intake_abs_status(rate: float | None) -> str:
    if rate is None or (isinstance(rate, float) and np.isnan(rate)):
        return NA_LABEL
    if rate >= INTAKE_ABS["HEALTHY"]: return "HEALTHY"
    if rate >= INTAKE_ABS["WATCH"]:   return "WATCH"
    return "AT RISK"


def booked_abs_status(rate: float | None) -> str:
    if rate is None or (isinstance(rate, float) and np.isnan(rate)):
        return NA_LABEL
    if rate >= BOOKED_ABS["HEALTHY"]: return "HEALTHY"
    if rate >= BOOKED_ABS["WATCH"]:   return "WATCH"
    return "AT RISK"


def m1_abs_status(rate: float | None) -> str:
    if rate is None or (isinstance(rate, float) and np.isnan(rate)):
        return NA_LABEL
    if rate >= M1_ABS["STRONG"]:   return "STRONG"
    if rate >= M1_ABS["MODERATE"]: return "MODERATE"
    return "LOW"


def fmt_pct(v, decimals=1) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    return f"{v:.{decimals}%}"


def fmt_pp(v) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    sign = "+" if v >= 0 else ""
    return f"{sign}{v:.1%}"


def fmt_num(v, decimals=2) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    return f"{v:.{decimals}f}"


def pct_rank(series: pd.Series) -> pd.Series:
    """Return percentile rank (0–100) for each element in a Series, ignoring NaN."""
    return series.rank(pct=True, na_option="keep") * 100


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════

def load_data() -> pd.DataFrame:
    print(f"Loading {DATA_PATH} …")
    df = pd.read_csv(DATA_PATH, low_memory=False)
    df["REFERRAL_DATE"] = pd.to_datetime(df["REFERRAL_DATE"], errors="coerce")
    df["month_of"] = df["REFERRAL_DATE"].dt.to_period("M").astype(str)

    # Northeast only
    df = df[df["AREA"] == "Northeast"].copy()

    # provider_id: prefer NPI, fall back to physician name
    npi = df["REFERRING_PROVIDER_NPI"].astype(str).str.strip()
    df["provider_id"] = np.where(
        (npi != "") & (npi.str.lower() != "nan"),
        npi,
        df["REFERRING_PHYSICIAN"],
    )

    # Conversion flags
    df["intake_started"] = (
        df["INTAKE_START_DATE"].notna() |
        df["APPOINTMENT_ID_FIRST_SCHEDULED"].notna()
    )
    df["visit_booked"] = df["APPOINTMENT_ID_FIRST_SCHEDULED"].notna()
    df["terminated"]   = df["TERMINATION_REASON"].notna()

    print(f"  {len(df):,} NE rows · {df['PARTNER_ASSIGNMENT'].nunique()} accounts")
    print(f"  Working days: Apr 1–{TODAY} = {APR_WORKDAYS}  |  Mar = {MAR_WORKDAYS}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# SIGNAL 1 — REFERRAL TREND
# ══════════════════════════════════════════════════════════════════════════════

def compute_referral_trend(df: pd.DataFrame, accounts: list[str]) -> pd.DataFrame:
    rows = []
    for acct in accounts:
        adf = df[df["PARTNER_ASSIGNMENT"] == acct]
        mar_refs = int((adf["month_of"] == LAST_MONTH).sum())
        apr_refs = int((adf["month_of"] == CURR_MONTH).sum())

        mar_rate = mar_refs / MAR_WORKDAYS if MAR_WORKDAYS > 0 else None
        apr_rate = apr_refs / APR_WORKDAYS if APR_WORKDAYS > 0 else None

        if mar_rate and mar_rate > 0:
            mom_pct = (apr_rate - mar_rate) / mar_rate
        else:
            mom_pct = None

        rows.append({
            "account":   acct,
            "mar_refs":  mar_refs,
            "apr_refs":  apr_refs,
            "mar_rpd":   mar_rate,        # refs per working day
            "apr_rpd":   apr_rate,
            "mom_pct":   mom_pct,
        })

    out = pd.DataFrame(rows).set_index("account")
    out["trend_pct_rank"] = pct_rank(out["mom_pct"])
    out["trend_status"]   = out["trend_pct_rank"].map(pct_status)
    return out


# ══════════════════════════════════════════════════════════════════════════════
# SIGNAL 2 — CONVERSION FUNNEL
# ══════════════════════════════════════════════════════════════════════════════

def _conv_rates(df: pd.DataFrame, month: str) -> dict[str, dict]:
    """Ref→Intake and Intake→Booked rates per account for one month.
    All referrals (including terminated) are included in both numerator and denominator.
    Terminated referrals that started intake still count toward the intake rate.
    """
    mdf = df[df["month_of"] == month]
    out = {}
    for acct, grp in mdf.groupby("PARTNER_ASSIGNMENT"):
        n = len(grp)
        if n < MIN_REFS:
            continue
        intake_n = int(grp["intake_started"].sum())
        out[acct] = {
            "refs":         n,
            "intake_rate":  intake_n / n,
            "booked_rate":  grp["visit_booked"].sum() / intake_n if intake_n > 0 else 0,
        }
    return out


def compute_conversion(df: pd.DataFrame, accounts: list[str]) -> pd.DataFrame:
    rates_last  = _conv_rates(df, LAST_MONTH)    # March
    rates_prior = _conv_rates(df, PRIOR_MONTH)   # February

    rows = []
    for acct in accounts:
        ml = rates_last.get(acct)
        mp = rates_prior.get(acct)

        ir_last  = ml["intake_rate"] if ml else None
        br_last  = ml["booked_rate"] if ml else None
        ir_prior = mp["intake_rate"] if mp else None
        br_prior = mp["booked_rate"] if mp else None

        ir_mom = (ir_last - ir_prior) if (ir_last is not None and ir_prior is not None) else None
        br_mom = (br_last - br_prior) if (br_last is not None and br_prior is not None) else None

        rows.append({
            "account":       acct,
            "refs_mar":      ml["refs"] if ml else None,
            "intake_rate":   ir_last,
            "booked_rate":   br_last,
            "intake_mom_pp": ir_mom,
            "booked_mom_pp": br_mom,
        })

    out = pd.DataFrame(rows).set_index("account")

    # Absolute status for current rates
    out["intake_status"] = out["intake_rate"].map(intake_abs_status)
    out["booked_status"] = out["booked_rate"].map(booked_abs_status)

    # Percentile status for MoM changes
    out["intake_mom_rank"]   = pct_rank(out["intake_mom_pp"])
    out["booked_mom_rank"]   = pct_rank(out["booked_mom_pp"])
    out["intake_mom_status"] = out["intake_mom_rank"].map(pct_status)
    out["booked_mom_status"] = out["booked_mom_rank"].map(pct_status)

    # Patch NA_LABEL for accounts with insufficient data
    for col in ["intake_status", "booked_status", "intake_mom_status", "booked_mom_status"]:
        out.loc[out["intake_rate"].isna() & (col in ["intake_status", "intake_mom_status"]), col] = NA_LABEL
        out.loc[out["booked_rate"].isna() & (col in ["booked_status", "booked_mom_status"]), col] = NA_LABEL

    return out


# ══════════════════════════════════════════════════════════════════════════════
# SIGNAL 3 — M1 RETENTION
# ══════════════════════════════════════════════════════════════════════════════

def _m1_rate(df_acct: pd.DataFrame, cohort_month: str, retention_month: str) -> float | None:
    """
    % of providers whose FIRST referral to this account was in cohort_month
    who also referred in retention_month.
    """
    first = (
        df_acct.groupby("provider_id")["REFERRAL_DATE"].min()
        .reset_index()
    )
    first["first_month"] = first["REFERRAL_DATE"].dt.to_period("M").astype(str)
    cohort = set(first[first["first_month"] == cohort_month]["provider_id"])

    if len(cohort) < MIN_COHORT:
        return None

    retained_ids = set(
        df_acct[df_acct["month_of"] == retention_month]["provider_id"].dropna()
    )
    return len(cohort & retained_ids) / len(cohort)


def compute_m1(df: pd.DataFrame, accounts: list[str]) -> pd.DataFrame:
    rows = []
    for acct in accounts:
        adf  = df[df["PARTNER_ASSIGNMENT"] == acct]
        curr = _m1_rate(adf, PRIOR_MONTH,  LAST_MONTH)    # Feb cohort → Mar
        prev = _m1_rate(adf, OLDEST_MONTH, PRIOR_MONTH)   # Jan cohort → Feb

        mom = (curr - prev) if (curr is not None and prev is not None) else None

        rows.append({
            "account":      acct,
            "m1_rate":      curr,
            "m1_prev_rate": prev,
            "m1_mom_pp":    mom,
        })

    out = pd.DataFrame(rows).set_index("account")
    out["m1_status"]      = out["m1_rate"].map(m1_abs_status)
    out["m1_mom_rank"]    = pct_rank(out["m1_mom_pp"])
    out["m1_mom_status"]  = out["m1_mom_rank"].map(pct_status)
    return out


# ══════════════════════════════════════════════════════════════════════════════
# SCORING
# ══════════════════════════════════════════════════════════════════════════════

# Point table — mirrors signal order left-to-right in the spreadsheet
POINT_TABLE: dict[str, dict[str, float]] = {
    # Signal 1 — Referral Trend (percentile-based)
    "trend_status": {
        "STRONG":               0,
        "GROWING":              0,
        "FLAT":                 5,
        "DECLINING":           15,
        "AT RISK":             30,
        NA_LABEL:               5,   # neutral = FLAT
    },
    # Signal 2A — Ref→Intake current (absolute thresholds)
    "intake_status": {
        "HEALTHY":              0,
        "WATCH":               15,
        "AT RISK":             30,
        NA_LABEL:               0,   # can't confirm a problem
    },
    # Signal 2A — Intake→Booked current (absolute thresholds)
    "booked_status": {
        "HEALTHY":              0,
        "WATCH":               15,
        "AT RISK":             30,
        NA_LABEL:               0,
    },
    # Signal 2B — Ref→Intake MoM (percentile-based)
    "intake_mom_status": {
        "STRONG":               0,
        "GROWING":              0,
        "FLAT":                 5,
        "DECLINING":           15,
        "AT RISK":             30,
        NA_LABEL:               5,   # neutral = FLAT
    },
    # Signal 2B — Intake→Booked MoM (percentile-based)
    "booked_mom_status": {
        "STRONG":               0,
        "GROWING":              0,
        "FLAT":                 5,
        "DECLINING":           15,
        "AT RISK":             30,
        NA_LABEL:               5,
    },
    # Signal 3A — M1 Retention current (absolute thresholds)
    "m1_status": {
        "STRONG":               0,
        "MODERATE":            15,
        "LOW":                 30,
        NA_LABEL:               0,
    },
    # Signal 3B — M1 Retention MoM (percentile-based)
    "m1_mom_status": {
        "STRONG":               0,
        "GROWING":              0,
        "FLAT":                 5,
        "DECLINING":           15,
        "AT RISK":             30,
        NA_LABEL:               5,   # neutral = FLAT
    },
}

MAX_RAW = 210   # 30×7 — worst-case sum across all signals

# Column labels for each scored signal (appear as score-section sub-headers)
SCORE_COL_LABELS = {
    "trend_status":       "Ref Trend\nPts",
    "intake_status":      "Ref→Intake\nPts",
    "booked_status":      "Intake→Booked\nPts",
    "intake_mom_status":  "Intake MoM\nPts",
    "booked_mom_status":  "Booked MoM\nPts",
    "m1_status":          "M1 Retention\nPts",
    "m1_mom_status":      "M1 MoM\nPts",
}


def compute_scores(master: pd.DataFrame, df: pd.DataFrame) -> pd.DataFrame:
    """Add per-signal point columns, raw score, volume multiplier, and final score."""

    # Per-signal points
    for sig, table in POINT_TABLE.items():
        pts_col = sig.replace("_status", "_pts")
        master[pts_col] = master[sig].map(lambda s, t=table: t.get(str(s), 0))

    # Raw score = sum of all point columns
    pts_cols = [s.replace("_status", "_pts") for s in POINT_TABLE]
    master["raw_score"] = master[pts_cols].sum(axis=1)

    # Volume multiplier = account's referral volume percentile within NE (min 0.1)
    vol = df.groupby("PARTNER_ASSIGNMENT")["REFERRAL_ID"].count()
    vol_pct = vol.rank(pct=True)
    master["vol_multiplier"] = vol_pct.clip(lower=0.1).round(2)

    # Opportunity Score — normalized to 0-100
    master["opp_score"] = (
        master["vol_multiplier"] * master["raw_score"] / MAX_RAW * 100
    ).round(1)

    return master


# ══════════════════════════════════════════════════════════════════════════════
# MASTER TABLE
# ══════════════════════════════════════════════════════════════════════════════

def build_master(df: pd.DataFrame) -> pd.DataFrame:
    accounts = sorted(df["PARTNER_ASSIGNMENT"].unique())

    trend = compute_referral_trend(df, accounts)
    conv  = compute_conversion(df, accounts)
    m1    = compute_m1(df, accounts)

    master = trend.join(conv, how="outer").join(m1, how="outer")

    # Add PPM
    ppm = df.groupby("PARTNER_ASSIGNMENT")["PPM"].first()
    master["PPM"] = ppm

    # Compute scores first, then sort by opportunity score desc
    master = compute_scores(master, df)

    vol = df.groupby("PARTNER_ASSIGNMENT")["REFERRAL_ID"].count()
    master["_vol"] = vol
    master = master.sort_values(["opp_score", "_vol"], ascending=[False, False]).drop(columns="_vol")

    master.index.name = "Account"
    return master


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITER  —  styled to match the reference Signals file
# ══════════════════════════════════════════════════════════════════════════════

# ── Exact colours from reference ───────────────────────────────────────────────
NAVY    = "1B2A4A"    # dark navy — all header backgrounds
WHITE   = "FFFFFF"
LGRAY   = "EFEFEF"    # section label background
ALT     = "F9F9F9"    # alternating row tint
DATA_FG = "333333"    # normal data text

# Status font colours (matching reference Definitions sheet exactly)
STATUS_FG = {
    "STRONG":    "2E7D32",
    "GROWING":   "558B2F",
    "HEALTHY":   "2E7D32",
    "FLAT":      "F57F17",
    "MODERATE":  "F57F17",
    "WATCH":     "F57F17",
    "DECLINING": "E65100",
    "LOW":       "C62828",
    "AT RISK":   "C62828",
}

# Subtle tint fills for status cells in the scorecard
STATUS_FILL = {
    "STRONG":    "E8F5E9",
    "GROWING":   "F1F8E9",
    "HEALTHY":   "E8F5E9",
    "FLAT":      "FFFDE7",
    "MODERATE":  "FFF8E1",
    "WATCH":     "FFF8E1",
    "DECLINING": "FBE9E7",
    "LOW":       "FFEBEE",
    "AT RISK":   "FFEBEE",
}

POS_FG  = "2E7D32"
NEG_FG  = "C62828"
BDR_CLR = "D0D0D0"


def _fill(hex_c: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_c)


def _font(bold=False, color=DATA_FG, size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _bdr() -> Border:
    s = Side(style="thin", color=BDR_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


def _col_hdr(ws, row, col, text, wrap=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
    c.fill      = _fill(NAVY)
    c.border    = _bdr()
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return c


def _section_label(ws, row, sc, ec, text):
    sl, el = get_column_letter(sc), get_column_letter(ec)
    if sc != ec:
        ws.merge_cells(f"{sl}{row}:{el}{row}")
    c = ws[f"{sl}{row}"]
    c.value     = text
    c.font      = Font(bold=True, size=10, color=NAVY, name="Calibri")
    c.fill      = _fill(LGRAY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18
    return c


def write_excel(master: pd.DataFrame):
    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1 — NE Account Signals  (scorecard)
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "NE Account Signals"
    ws.freeze_panes = "C6"    # freeze account + PPM cols, keep header rows

    # ── Column spec ────────────────────────────────────────────────────────────
    # (data_key, header_text, col_width, is_status, is_delta, is_pts, is_score)
    COLS = [
        # identifiers
        ("_acct",             "Account",                       32, False, False, False, False),
        ("PPM",               "PPM",                           18, False, False, False, False),
        # Signal 1 — Referral Trend
        ("mar_rpd",           f"Mar\nRefs/Day\n({MAR_WORKDAYS} days)",  11, False, False, False, False),
        ("apr_rpd",           f"Apr Refs/Day\n(thru {TODAY.strftime('%-m/%-d')},\n{APR_WORKDAYS} days)",
                                                               13, False, False, False, False),
        ("mom_pct",           "MoM %",                          9, False, True,  False, False),
        ("trend_status",      "Status",                        11, True,  False, False, False),
        # Signal 2A — Conversion Current
        ("refs_mar",          "Mar Refs",                       9, False, False, False, False),
        ("intake_rate",       "Ref→Intake\n(Mar)",             12, False, False, False, False),
        ("intake_status",     "Status",                        11, True,  False, False, False),
        ("booked_rate",       "Intake→Booked\n(Mar)",          13, False, False, False, False),
        ("booked_status",     "Status",                        11, True,  False, False, False),
        # Signal 2B — Conversion MoM
        ("intake_mom_pp",     "Ref→Intake\nMoM Δ",            12, False, True,  False, False),
        ("intake_mom_status", "Status",                        11, True,  False, False, False),
        ("booked_mom_pp",     "Intake→Booked\nMoM Δ",         13, False, True,  False, False),
        ("booked_mom_status", "Status",                        11, True,  False, False, False),
        # Signal 3A — M1 Current
        ("m1_rate",           "M1 Rate\n(Feb→Mar)",            12, False, False, False, False),
        ("m1_status",         "Status",                        11, True,  False, False, False),
        # Signal 3B — M1 MoM
        ("m1_mom_pp",         "M1 MoM Δ",                     10, False, True,  False, False),
        ("m1_mom_status",     "Status",                        11, True,  False, False, False),
        # ── Scoring section ──────────────────────────────────────────────────
        # Points columns mirror signal order exactly (left → right)
        ("trend_pts",         "Ref Trend\nPts",                 9, False, False, True,  False),
        ("intake_pts",        "Ref→Intake\nPts",               10, False, False, True,  False),
        ("booked_pts",        "Intake→Booked\nPts",            12, False, False, True,  False),
        ("intake_mom_pts",    "Intake MoM\nPts",               10, False, False, True,  False),
        ("booked_mom_pts",    "Booked MoM\nPts",               10, False, False, True,  False),
        ("m1_pts",            "M1 Retention\nPts",             11, False, False, True,  False),
        ("m1_mom_pts",        "M1 MoM\nPts",                    9, False, False, True,  False),
        # Totals
        ("raw_score",         f"Raw\nScore\n(max {MAX_RAW})",   9, False, False, False, True ),
        ("vol_multiplier",    "Volume\nMultiplier",              9, False, False, False, True ),
        ("opp_score",         "OPPORTUNITY\nSCORE\n(0–100)",   13, False, False, False, True ),
    ]
    n_cols   = len(COLS)
    # column index (1-based) where scoring section starts
    SCORE_START_COL = next(i+1 for i, (k,*_) in enumerate(COLS) if k == "trend_pts")

    # ── Row 1: Title ───────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value     = "Northeast Region — Account Signals"
    c.font      = Font(bold=True, size=14, color=NAVY, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Row 2: Data updated as of ──────────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    c = ws["A2"]
    c.value     = f"Data updated as of:  {TODAY.strftime('%B %d, %Y')}"
    c.font      = Font(size=10, color="666666", name="Calibri")
    c.fill      = _fill("F3F3F3")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Row 3: blank ───────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Row 4: Section band labels ─────────────────────────────────────────────
    SECTIONS = [
        (1,  2,  ""),
        (3,  6,  f"REFERRAL TREND  —  Apr 1–{TODAY.strftime('%-m/%-d')} vs full March  (refs per working day)"),
        (7,  11, f"CONVERSION FUNNEL — CURRENT  (March, all referrals included)"),
        (12, 15, f"CONVERSION FUNNEL — MoM  (March vs February, pp change, percentile ranked)"),
        (16, 17, f"M1 RETENTION — CURRENT  (Feb-new providers who referred again in March)"),
        (18, 19, f"M1 RETENTION — MoM  (Feb cohort vs Jan cohort, pp change, percentile ranked)"),
        (SCORE_START_COL, n_cols,
                 f"OPPORTUNITY SCORE  (0–100)  ·  Score = Volume Multiplier × Raw Points / {MAX_RAW} × 100"
                 f"  ·  Points mirror signals left-to-right  ·  Higher = bigger opportunity"),
    ]
    for sc, ec, label in SECTIONS:
        _section_label(ws, 4, sc, ec, label)

    # Score section gets a distinct darker background to visually separate it
    score_label_cell = ws[f"{get_column_letter(SCORE_START_COL)}4"]
    score_label_cell.fill = _fill(NAVY)
    score_label_cell.font = Font(bold=True, size=10, color=WHITE, name="Calibri")

    # ── Row 5: Column headers ──────────────────────────────────────────────────
    for col_idx, (key, hdr, width, is_status, is_delta, is_pts, is_score) in enumerate(COLS, start=1):
        c = _col_hdr(ws, 5, col_idx, hdr)
        # Score section headers get a slightly different shade to match section band
        if is_pts:
            c.fill = _fill("243757")   # slightly lighter navy for pts columns
        if is_score:
            c.fill = _fill(NAVY)
            c.font = Font(bold=True, size=10, color="FFD700", name="Calibri")  # gold for score totals
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[5].height = 40

    # ── Rows 6+: Data ─────────────────────────────────────────────────────────
    for row_idx, (acct, row) in enumerate(master.iterrows(), start=6):
        alt = (row_idx % 2 == 0)
        row_bg = ALT if alt else WHITE

        def _raw(key):
            return acct if key == "_acct" else row.get(key)

        def _fmt(key, val):
            if val is None or (isinstance(val, float) and np.isnan(val)):
                return "—"
            if key in ("mar_rpd", "apr_rpd"):
                return fmt_num(val, 2)
            if key in ("mom_pct", "intake_mom_pp", "booked_mom_pp", "m1_mom_pp"):
                return fmt_pp(val)
            if key in ("intake_rate", "booked_rate", "m1_rate"):
                return fmt_pct(val)
            if key == "refs_mar":
                return f"{int(val):,}"
            if key == "vol_multiplier":
                return fmt_num(val, 2)
            if key == "opp_score":
                return round(val, 1) if val is not None else "—"
            if key in ("raw_score",) or key.endswith("_pts"):
                return int(val) if val is not None else 0
            return val

        for col_idx, (key, hdr, width, is_status, is_delta, is_pts, is_score) in enumerate(COLS, start=1):
            raw     = _raw(key)
            display = raw if is_status else _fmt(key, raw)

            c = ws.cell(row=row_idx, column=col_idx, value=display)
            c.border    = _bdr()
            c.alignment = Alignment(
                horizontal="left"   if col_idx <= 2 else "center",
                vertical="center",
            )

            status_str = str(raw) if raw is not None else ""

            if is_status:
                # Tinted background + coloured bold text matching reference style
                fill_hex = STATUS_FILL.get(status_str, "F3F3F3")
                fg_hex   = STATUS_FG.get(status_str, DATA_FG)
                c.fill = _fill(fill_hex)
                c.font = Font(bold=True, size=10, color=fg_hex, name="Calibri")

            elif is_pts:
                # Points cells: light blue-gray tint, pts value in navy
                pts_bg = "EEF2F7" if not alt else "E4EAF2"
                c.fill = _fill(pts_bg)
                pts_val = raw if raw is not None else 0
                # Colour the number: red if points > 0 (problem), gray if 0 (no issue)
                if isinstance(pts_val, (int, float)) and pts_val > 0:
                    c.font = Font(bold=True, size=10, color=NEG_FG, name="Calibri")
                else:
                    c.font = Font(size=10, color="AAAAAA", name="Calibri")

            elif is_score:
                # Final score columns: dark background, gold text for the score
                if key == "opp_score":
                    c.fill = _fill(NAVY)
                    c.font = Font(bold=True, size=11, color="FFD700", name="Calibri")
                elif key == "raw_score":
                    raw_bg = "1F3560" if not alt else "1A2D52"
                    c.fill = _fill(raw_bg)
                    c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
                else:  # vol_multiplier
                    raw_bg = "253E6E" if not alt else "1F3560"
                    c.fill = _fill(raw_bg)
                    c.font = Font(size=10, color="C8D8F0", name="Calibri")

            elif is_delta and display not in ("—", None, ""):
                c.fill = _fill(row_bg)
                try:
                    num = float(str(display).replace("+", "").replace("%", "")) / 100
                    if num > 0.001:
                        c.font = Font(bold=True, size=10, color=POS_FG, name="Calibri")
                    elif num < -0.001:
                        c.font = Font(bold=True, size=10, color=NEG_FG, name="Calibri")
                    else:
                        c.font = _font(size=10)
                except ValueError:
                    c.font = _font(size=10)

            else:
                c.fill = _fill(row_bg)
                c.font = _font(bold=(col_idx <= 2), size=10)

        ws.row_dimensions[row_idx].height = 16

    # ── Footnote row ───────────────────────────────────────────────────────────
    foot_row = 6 + len(master)
    ws.merge_cells(f"A{foot_row}:{get_column_letter(n_cols)}{foot_row}")
    c = ws[f"A{foot_row}"]
    c.value = (
        "Conversion rates include all referrals (terminated referrals that started intake count toward the intake rate).  "
        "'Not enough referrals' = fewer than 5 refs in month (conversion) or fewer than 3 new providers in cohort (M1)."
    )
    c.font      = Font(italic=True, size=9, color="666666", name="Calibri")
    c.fill      = _fill("F3F3F3")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[foot_row].height = 14

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2 — Definitions  (matching reference Definitions sheet style)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Definitions")
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 52
    ws2.column_dimensions["D"].width = 32

    r = 1
    # Title
    ws2.merge_cells(f"A{r}:D{r}")
    c = ws2[f"A{r}"]
    c.value     = "Area Assessment Logic"
    c.font      = Font(bold=True, size=16, color=NAVY, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[r].height = 28
    r += 1

    ws2.merge_cells(f"A{r}:D{r}")
    c = ws2[f"A{r}"]
    c.value = "This guide explains every status and threshold used in the NE Account Signals scorecard."
    c.font  = Font(size=10, color="666666", name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[r].height = 16
    r += 2

    def _d_section(row, text):
        ws2.merge_cells(f"A{row}:D{row}")
        c = ws2[f"A{row}"]
        c.value     = text
        c.font      = Font(bold=True, size=12, color=NAVY, name="Calibri")
        c.fill      = _fill(LGRAY)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[row].height = 20

    def _d_hdr(row, *vals):
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=row, column=ci, value=v)
            c.font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
            c.fill      = _fill(NAVY)
            c.border    = _bdr()
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[row].height = 16

    def _d_row(row, status_label, col2, col3, col4="", alt_row=False):
        bg = ALT if alt_row else WHITE
        vals = [status_label, col2, col3, col4]
        fg = STATUS_FG.get(status_label, DATA_FG)
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=row, column=ci, value=v)
            c.border    = _bdr()
            c.alignment = Alignment(horizontal="center" if ci == 1 else "left",
                                    vertical="center")
            c.fill = _fill(bg)
            if ci == 1:
                c.font = Font(bold=True, size=9, color=fg, name="Calibri")
            else:
                c.font = Font(size=9, color=DATA_FG, name="Calibri")
        ws2.row_dimensions[row].height = 14

    # — Referral Trend ─────────────────────────────────────────────────────────
    _d_section(r, "REFERRAL TREND  —  MoM % change in refs per working day (Apr vs Mar)"); r += 1
    _d_hdr(r, "Status", "Percentile Threshold", "What It Means", "What to Do"); r += 1
    for i, (s, pct, meaning, action) in enumerate([
        ("STRONG",    "≥ 80th pct", "Top 20% — growing faster than most NE accounts", "Protect & expand"),
        ("GROWING",   "≥ 60th pct", "Above average — positive momentum",               "Support growth"),
        ("FLAT",      "≥ 40th pct", "Middle of the pack — no clear direction",         "Investigate why growth stalled"),
        ("DECLINING", "≥ 20th pct", "Below average — volume dropping",                 "Root cause — build a 2-week plan"),
        ("AT RISK",   "< 20th pct", "Bottom of the portfolio — steepest decline",      "Immediate escalation"),
    ]):
        _d_row(r, s, pct, meaning, action, alt_row=(i % 2 == 1)); r += 1
    r += 1

    # — Conversion Current ─────────────────────────────────────────────────────
    _d_section(r, "CONVERSION FUNNEL — CURRENT  (March rates, all referrals included)"); r += 1
    _d_hdr(r, "Stage", "Healthy ≥", "Watch ≥", "At Risk <"); r += 1
    for i, (stage, h, w, ar) in enumerate([
        ("Ref → Intake",    "55%", "45%", "45%"),
        ("Intake → Booked", "70%", "60%", "60%"),
    ]):
        bg = ALT if i % 2 == 1 else WHITE
        for ci, v in enumerate([stage, h, w, ar], 1):
            c = ws2.cell(row=r, column=ci, value=v)
            c.font   = Font(size=9, color=DATA_FG, name="Calibri")
            c.border = _bdr()
            c.fill   = _fill(bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r].height = 14; r += 1
    ws2.merge_cells(f"A{r}:D{r}")
    ws2[f"A{r}"].value = "Key consideration: minimum 5 referrals in month required. Below 5 = 'Not enough referrals'."
    ws2[f"A{r}"].font  = Font(italic=True, size=9, color="666666", name="Calibri")
    ws2.row_dimensions[r].height = 13; r += 2

    # — Conversion MoM ─────────────────────────────────────────────────────────
    _d_section(r, "CONVERSION FUNNEL — MoM  (March vs February pp change, percentile ranked within NE)"); r += 1
    _d_hdr(r, "Status", "Percentile Threshold", "Applied to both:", ""); r += 1
    for i, (s, pct, meaning) in enumerate([
        ("STRONG",    "≥ 80th pct", "Ref→Intake MoM Δ  and  Intake→Booked MoM Δ  (ranked separately)"),
        ("GROWING",   "≥ 60th pct", ""),
        ("FLAT",      "≥ 40th pct", ""),
        ("DECLINING", "≥ 20th pct", ""),
        ("AT RISK",   "< 20th pct", ""),
    ]):
        _d_row(r, s, pct, meaning, "", alt_row=(i % 2 == 1)); r += 1
    r += 1

    # — M1 Current ─────────────────────────────────────────────────────────────
    _d_section(r, "M1 RETENTION — CURRENT  (% of Feb-new providers who also referred in March)"); r += 1
    _d_hdr(r, "Status", "Threshold", "What It Means", "Action"); r += 1
    for i, (s, thr, meaning, action) in enumerate([
        ("STRONG",   "≥ 35%", "≥ 35% of new providers come back — strong retention",  "Protect these relationships"),
        ("MODERATE", "≥ 25%", "25–35% return — some drop-off, worth monitoring",      "Investigate drop-offs"),
        ("LOW",      "< 25%", "< 25% — most refer once and stop",                     "Fix physician experience first"),
    ]):
        _d_row(r, s, thr, meaning, action, alt_row=(i % 2 == 1)); r += 1
    ws2.merge_cells(f"A{r}:D{r}")
    ws2[f"A{r}"].value = "Cohort = providers whose first-ever referral to this account was in the cohort month. Minimum 3 new providers required."
    ws2[f"A{r}"].font  = Font(italic=True, size=9, color="666666", name="Calibri")
    ws2.row_dimensions[r].height = 13; r += 2

    # — M1 MoM ─────────────────────────────────────────────────────────────────
    _d_section(r, "M1 RETENTION — MoM  (Feb cohort retention vs Jan cohort retention, percentile ranked)"); r += 1
    _d_hdr(r, "Status", "Percentile Threshold", "What It Means", ""); r += 1
    for i, (s, pct, meaning) in enumerate([
        ("STRONG",    "≥ 80th pct", "Retention improving meaningfully vs prior cohort"),
        ("GROWING",   "≥ 60th pct", "Slight improvement"),
        ("FLAT",      "≥ 40th pct", "No significant change"),
        ("DECLINING", "≥ 20th pct", "Retention slipping"),
        ("AT RISK",   "< 20th pct", "Retention significantly worse than prior cohort"),
    ]):
        _d_row(r, s, pct, meaning, "", alt_row=(i % 2 == 1)); r += 1

    # ── Opportunity Score table ────────────────────────────────────────────────
    r += 1
    ws2.merge_cells(f"A{r}:D{r}")
    c = ws2[f"A{r}"]
    c.value     = "OPPORTUNITY SCORE  (0–100)"
    c.font      = Font(bold=True, size=16, color=NAVY, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[r].height = 28; r += 1

    ws2.merge_cells(f"A{r}:D{r}")
    c = ws2[f"A{r}"]
    c.value = (
        f"Ranks accounts by how much upside there is if you fix the problems. "
        f"Higher score = bigger opportunity.  "
        f"Formula: Score = Volume Multiplier × Raw Points ÷ {MAX_RAW} × 100"
    )
    c.font  = Font(size=10, color="666666", name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[r].height = 16; r += 1

    ws2.merge_cells(f"A{r}:D{r}")
    c = ws2[f"A{r}"]
    c.value = "Volume Multiplier = account's referral volume percentile within NE (min 0.1). Larger accounts with same problems score higher."
    c.font  = Font(italic=True, size=9, color="666666", name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[r].height = 13; r += 2

    # Header row
    _d_section(r, "OPPORTUNITY SCORE — Point Table"); r += 1
    for ci, v in enumerate(["Component", "Status", "Points", "Rationale"], 1):
        c = ws2.cell(row=r, column=ci, value=v)
        c.font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
        c.fill      = _fill(NAVY)
        c.border    = _bdr()
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[r].height = 16; r += 1

    # Score rows
    score_rows = [
        # component, status, points, rationale
        ("Multiplier Floor",       "—",              "0.1×",  "Volume multiplier minimum — even tiny accounts are scored"),
        ("",                       "",               "",       ""),
        ("Ref → Intake",           "AT RISK",        25,      "Large conversion gap — biggest upside"),
        ("Ref → Intake",           "WATCH",          15,      "Moderate gap"),
        ("Ref → Intake",           "HEALTHY",         0,      "No action needed"),
        ("",                       "",               "",       ""),
        ("Intake → Booked",        "AT RISK",        25,      "Large booking gap"),
        ("Intake → Booked",        "WATCH",          15,      "Moderate gap"),
        ("Intake → Booked",        "HEALTHY",         0,      "No action needed"),
        ("",                       "",               "",       ""),
        ("M1 Retention",           "LOW",            20,      "Most new providers don't return — high recovery upside"),
        ("M1 Retention",           "MODERATE",        8,      "Some drop-off"),
        ("M1 Retention",           "STRONG",          0,      "No action needed"),
        ("",                       "",               "",       ""),
        ("Referral Trend",         "DECLINING",      20,      "Actively declining — recoverable upside (was healthy before)"),
        ("Referral Trend",         "AT RISK",        15,      "Bottom of portfolio — may already be near floor"),
        ("Referral Trend",         "FLAT",            5,      "No growth — soft opportunity"),
        ("Referral Trend",         "GROWING",         0,      "Positive momentum — no action needed"),
        ("Referral Trend",         "STRONG",          0,      "No action needed"),
        ("",                       "",               "",       ""),
        ("Ref → Intake MoM",       "DECLINING",      10,      "Conversion declining — adds urgency to current state"),
        ("Ref → Intake MoM",       "AT RISK",         8,      "Conversion nosediving"),
        ("Ref → Intake MoM",       "FLAT",            2,      "Not recovering on its own"),
        ("Ref → Intake MoM",       "GROWING/STRONG",  0,      "Improving — no action needed"),
        ("",                       "",               "",       ""),
        ("Intake → Booked MoM",    "DECLINING",      10,      "Booking rate declining"),
        ("Intake → Booked MoM",    "AT RISK",         8,      "Booking rate nosediving"),
        ("Intake → Booked MoM",    "FLAT",            2,      "Not recovering on its own"),
        ("Intake → Booked MoM",    "GROWING/STRONG",  0,      "Improving — no action needed"),
        ("",                       "",               "",       ""),
        ("M1 Retention MoM",       "DECLINING",       8,      "Retention worsening vs prior cohort"),
        ("M1 Retention MoM",       "AT RISK",         5,      "Retention significantly worse"),
        ("M1 Retention MoM",       "FLAT",            1,      "Not improving"),
        ("M1 Retention MoM",       "GROWING/STRONG",  0,      "Improving — no action needed"),
        ("",                       "",               "",       ""),
        ("No-Data Factor",         "Missing signal", "FLAT pts", "Insufficient data → neutral score, not 0 or max"),
    ]

    alt_i = 0
    for comp, status, pts, rationale in score_rows:
        if comp == "":
            ws2.row_dimensions[r].height = 5; r += 1
            continue
        bg = ALT if alt_i % 2 == 1 else WHITE
        alt_i += 1
        fg_status = STATUS_FG.get(status, DATA_FG)
        row_vals = [comp, status, pts, rationale]
        for ci, val in enumerate(row_vals, 1):
            c = ws2.cell(row=r, column=ci, value=val)
            c.border    = _bdr()
            c.fill      = _fill(bg)
            c.alignment = Alignment(horizontal="center" if ci in (2, 3) else "left",
                                    vertical="center")
            if ci == 2 and status in STATUS_FG:
                c.font = Font(bold=True, size=9, color=fg_status, name="Calibri")
            elif ci == 3:
                pts_num = pts if isinstance(pts, (int, float)) else 0
                pts_color = NEG_FG if isinstance(pts_num, (int, float)) and pts_num > 0 else DATA_FG
                c.font = Font(bold=True, size=9, color=pts_color, name="Calibri")
            else:
                c.font = Font(size=9, color=DATA_FG, name="Calibri", bold=(ci == 1))
        ws2.row_dimensions[r].height = 14; r += 1

    # Max score footnote
    r += 1
    ws2.merge_cells(f"A{r}:D{r}")
    c = ws2[f"A{r}"]
    c.value = f"Maximum raw score = {MAX_RAW} pts (25+25+20+20+10+10+8). With volume multiplier = 1.0, max Opportunity Score = 100."
    c.font  = Font(italic=True, size=9, color="666666", name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[r].height = 13

    wb.save(OUTPUT_PATH)
    print(f"\n✅  Saved → {OUTPUT_PATH}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    df     = load_data()
    master = build_master(df)

    print(f"\n── Preview (top 8 accounts) ──")
    preview_cols = [
        "PPM", "mar_refs", "apr_refs", "mom_pct", "trend_status",
        "intake_rate", "intake_status", "booked_rate", "booked_status",
        "intake_mom_pp", "intake_mom_status", "booked_mom_pp", "booked_mom_status",
        "m1_rate", "m1_status", "m1_mom_pp", "m1_mom_status",
    ]
    with pd.option_context("display.max_columns", 20, "display.width", 200):
        print(master[preview_cols].head(8).to_string())

    print("\n── Status distributions ──")
    for col in ["trend_status", "intake_status", "booked_status",
                "intake_mom_status", "booked_mom_status", "m1_status", "m1_mom_status"]:
        print(f"\n{col}:")
        print(master[col].value_counts().to_string())

    write_excel(master)


if __name__ == "__main__":
    main()
